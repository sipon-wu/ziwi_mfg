# Excel 导入服务
import io
import os
import uuid
import json
from typing import Optional, List
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.repositories.excel_import_repo import ExcelImportRepository

# ========= 导入模板定义 =========
TEMPLATES = {
    "production": {
        "name": "生产日报模板",
        "description": "用于导入生产日报数据（产线、产品型号、计划产量、实际产量、不良数、工时）",
        "columns": [
            {"field": "report_date", "label": "日期", "required": True, "type": "date", "format": "YYYY-MM-DD"},
            {"field": "line_code", "label": "产线编码", "required": True, "type": "string"},
            {"field": "product_code", "label": "产品编码", "required": True, "type": "string"},
            {"field": "planned_qty", "label": "计划产量", "required": True, "type": "number", "min": 0},
            {"field": "actual_qty", "label": "实际产量", "required": True, "type": "number", "min": 0},
            {"field": "defect_qty", "label": "不良数", "required": False, "type": "number", "min": 0, "default": 0},
            {"field": "work_hours", "label": "工时(小时)", "required": False, "type": "number", "min": 0},
        ]
    },
    "tpm": {
        "name": "设备点检模板",
        "description": "用于导入设备点检记录",
        "columns": [
            {"field": "check_date", "label": "日期", "required": True, "type": "date"},
            {"field": "equipment_code", "label": "设备编码", "required": True, "type": "string"},
            {"field": "check_item", "label": "点检项目", "required": True, "type": "string"},
            {"field": "result", "label": "结果(正常/异常)", "required": True, "type": "string", "enum": ["正常", "异常"]},
            {"field": "remark", "label": "异常描述", "required": False, "type": "string"},
        ]
    },
    "quality": {
        "name": "检验记录模板",
        "description": "用于导入品质检验记录",
        "columns": [
            {"field": "inspection_date", "label": "检验日期", "required": True, "type": "date"},
            {"field": "product_code", "label": "产品编码", "required": True, "type": "string"},
            {"field": "inspection_type", "label": "检验类型(首检/巡检/抽检)", "required": True, "type": "string"},
            {"field": "sample_size", "label": "样本数", "required": True, "type": "number", "min": 1},
            {"field": "defect_qty", "label": "不良数", "required": True, "type": "number", "min": 0},
            {"field": "decision", "label": "判定(合格/不合格/让步)", "required": True, "type": "string"},
        ]
    },
    "andon": {
        "name": "安灯记录模板",
        "description": "用于导入安灯呼叫记录",
        "columns": [
            {"field": "call_date", "label": "日期", "required": True, "type": "date"},
            {"field": "call_type", "label": "异常类型", "required": True, "type": "string"},
            {"field": "station", "label": "工位", "required": True, "type": "string"},
            {"field": "description", "label": "异常描述", "required": True, "type": "string"},
            {"field": "response_minutes", "label": "响应时长(分钟)", "required": False, "type": "number"},
        ]
    },
    "energy": {
        "name": "能耗日更模板",
        "description": "用于导入能耗数据",
        "columns": [
            {"field": "record_date", "label": "日期", "required": True, "type": "date"},
            {"field": "meter_code", "label": "计量表编码", "required": True, "type": "string"},
            {"field": "reading", "label": "读数", "required": True, "type": "number"},
            {"field": "multiplier", "label": "倍率", "required": False, "type": "number", "default": 1},
            {"field": "usage", "label": "用量", "required": True, "type": "number"},
        ]
    },
    "employee": {
        "name": "人员出勤模板",
        "description": "用于导入人员出勤数据",
        "columns": [
            {"field": "date", "label": "日期", "required": True, "type": "date"},
            {"field": "employee_no", "label": "工号", "required": True, "type": "string"},
            {"field": "attendance_status", "label": "出勤状态(在岗/休息/请假/出差)", "required": True, "type": "string"},
            {"field": "team_name", "label": "班组", "required": True, "type": "string"},
        ]
    },
    "equipment": {
        "name": "设备运行记录模板",
        "description": "用于导入设备运行数据",
        "columns": [
            {"field": "record_date", "label": "日期", "required": True, "type": "date"},
            {"field": "equipment_code", "label": "设备编码", "required": True, "type": "string"},
            {"field": "start_time", "label": "开机时间", "required": True, "type": "datetime"},
            {"field": "end_time", "label": "关机时间", "required": True, "type": "datetime"},
            {"field": "running_hours", "label": "运行时长(小时)", "required": True, "type": "number"},
        ]
    },
    "inventory": {
        "name": "库存导入模板",
        "description": "用于导入库存数据",
        "columns": [
            {"field": "material_code", "label": "物料编码", "required": True, "type": "string"},
            {"field": "material_name", "label": "物料名称", "required": True, "type": "string"},
            {"field": "quantity", "label": "数量", "required": True, "type": "number", "min": 0},
            {"field": "unit", "label": "单位", "required": True, "type": "string"},
            {"field": "warehouse", "label": "仓库", "required": True, "type": "string"},
        ]
    },
}


class ExcelImportService:
    def __init__(self, repo: ExcelImportRepository):
        self.repo = repo

    async def list_templates(self) -> list:
        templates = []
        for ttype, tpl in TEMPLATES.items():
            templates.append({
                "import_type": ttype,
                "template_name": tpl["name"],
                "description": tpl.get("description", ""),
                "columns": tpl["columns"],
                "version": "1.0",
            })
        return templates

    def generate_template_file(self, import_type: str) -> bytes:
        """生成 Excel 模板文件字节流"""
        tpl = TEMPLATES.get(import_type)
        if not tpl:
            raise HTTPException(status_code=404, detail={"code": "404-0000", "message": f"不支持的导入类型: {import_type}"})

        wb = Workbook()
        ws = wb.active
        ws.title = tpl["name"]

        # 样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0D7377", end_color="0D7377", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # 写入表头
        headers = [c["label"] for c in tpl["columns"]]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # 第2行写入示例
        for col_idx, col in enumerate(tpl["columns"], 1):
            sample = f"示例{col['label']}"
            ws.cell(row=2, column=col_idx, value=sample).border = thin_border

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def parse_and_validate(self, import_type: str, file_bytes: bytes) -> dict:
        """解析 Excel 并校验数据"""
        tpl = TEMPLATES.get(import_type)
        if not tpl:
            raise HTTPException(status_code=400, detail={"code": "400-0000", "message": f"不支持的导入类型: {import_type}"})

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        rows = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if all(v is None for v in row):
                continue  # 跳过空行
            row_data = {}
            row_errors = []
            for col_idx, col in enumerate(tpl["columns"]):
                val = row[col_idx] if col_idx < len(row) else None
                field = col["field"]

                # 必填校验
                if col.get("required") and (val is None or str(val).strip() == ""):
                    row_errors.append(f"{col['label']} 为必填项")
                    continue

                if val is None:
                    val = col.get("default")
                    if val is None:
                        continue

                # 类型校验
                if col["type"] == "number":
                    try:
                        val = float(val)
                        if col.get("min") is not None and val < col["min"]:
                            row_errors.append(f"{col['label']} 不能小于 {col['min']}")
                    except (ValueError, TypeError):
                        row_errors.append(f"{col['label']} 必须为数字")

                # 枚举校验
                if col.get("enum") and str(val) not in col["enum"]:
                    row_errors.append(f"{col['label']} 必须在 [{','.join(col['enum'])}] 范围内")

                row_data[field] = val

            if row_errors:
                errors.append({"row": row_idx, "errors": row_errors})
            else:
                rows.append(row_data)

        return {"rows": rows, "errors": errors, "total": len(rows) + len(errors)}

    async def create_import_task(self, data: dict) -> dict:
        task_id = await self.repo.create_task(data)
        return {"task_id": task_id, "status": "pending"}

    async def get_task(self, task_id: int) -> Optional[dict]:
        task = await self.repo.get_task(task_id)
        if not task:
            raise HTTPException(status_code=404, detail={"code": "404-0000", "message": "导入任务不存在"})
        return task

    async def list_tasks(self, page: int = 1, page_size: int = 20, import_type: str = None, status: str = None) -> dict:
        return await self.repo.list_tasks(page, page_size, import_type, status)
